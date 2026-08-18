"""apply_sku_map_master.py — sync canonical SKUs to sku_map_MASTER.xlsx (LIVE DB).

The dashboard resolves canonical SKUs from Postgres (the `cogs_aliases` table — the live
form of sku_aliases.csv), NOT from the spreadsheet, so editing the xlsx alone changes
nothing. This tool closes that gap:

  1. imports the sheet's alias->canonical pairs into cogs_aliases (upsert; idempotent);
  2. re-resolves EVERY pl_line_items row's canonical through the refreshed alias table,
     via pl_tracker.reprocess_from_stored_events() (rebuilds from the stored ledger, NO
     network) — this is the step that makes historic rows change canonical;
  3. the /pl rollup then merges split rows automatically, because it GROUPs BY
     canonical_sku (e.g. `6169` + `BD-6169` collapse into `BD-6169`).

COLOUR is kept distinct by default: any pair whose alias colour differs from its canonical
colour (e.g. v-plo+Pk -> v-plo+Gy) is SKIPPED so colours never collapse into one another;
same-colour and non-colour normalisations (6337-P2 -> BD-6337-P2) still apply. The dry-run
lists every skipped pair. Pass --allow-colour-merge to disable the guard (not recommended).

DRY-RUN by default: reads the sheet + DB and reports (a) the colour-crossing pairs skipped,
(b) the alias upsert delta, and (c) exactly which rows would change canonical and how many
distinct canonicals collapse. Writes NOTHING.
  --apply : upserts the aliases, then runs the full reprocess. TAKE A DB BACKUP FIRST.

Targets whatever db.connect() points at — set DATABASE_URL for the LIVE Railway DB:
    set DATABASE_URL=postgresql://...
    set SKU_MAP_XLSX=sku_map_MASTER.xlsx      (optional; default this name in cwd)
    python apply_sku_map_master.py            # dry-run report
    python apply_sku_map_master.py --apply    # after a DB backup

The first printed line states Postgres (LIVE) vs local SQLite, so you can confirm it ran
against the right database.
"""
import os
import sys
import html
from datetime import datetime, timezone

import db
import pl_cogs  # noqa: F401 -- ensures schema module importable; resolver logic mirrored below

SHEET = "Map"
CANON_HDR = "CANONICAL SKU (going to be used )"   # NB: trailing space is intentional
ALIAS_HDRS = ["mapped sku 1", "mapped sku 2", "mapped sku 3",
              "mapped sku 4", "mapped sku 5", "mapped sku 6"]


def _now():
    return datetime.now(timezone.utc).isoformat()


def load_pairs(xlsx_path):
    """Return (ordered_pairs, n_data_rows). ordered_pairs = [(variant_sku, canonical)]
    in sheet order, one per non-empty 'mapped sku N' cell."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed — run: pip install openpyxl")
    if not os.path.exists(xlsx_path):
        sys.exit(f"spreadsheet not found: {xlsx_path} (set SKU_MAP_XLSX or run from its folder)")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"sheet {SHEET!r} not found; sheets = {wb.sheetnames}")
    it = wb[SHEET].iter_rows(values_only=True)
    header = list(next(it))
    idx = {h: i for i, h in enumerate(header)}
    if CANON_HDR not in idx:
        sys.exit(f"canonical column {CANON_HDR!r} not found (check the trailing space!). "
                 f"Headers: {header}")
    ci = idx[CANON_HDR]
    ai = [idx[h] for h in ALIAS_HDRS if h in idx]
    if not ai:
        sys.exit(f"no 'mapped sku N' columns found. Headers: {header}")
    ordered, n_rows = [], 0
    for r in it:
        canon = r[ci] if len(r) > ci else None
        if canon is None or str(canon).strip() == "":
            continue
        canon = str(canon).strip()
        n_rows += 1
        for j in ai:
            v = r[j] if len(r) > j else None
            if v is not None and str(v).strip() != "":
                ordered.append((str(v).strip(), canon))
    return ordered, n_rows


def dedup(ordered):
    """Collapse to one canonical per exact variant_sku (LAST occurrence wins), and report
    any variant_sku the sheet maps to more than one canonical."""
    seen, conflicts = {}, {}
    for v, c in ordered:
        if v in seen and seen[v] != c:
            conflicts.setdefault(v, {seen[v]}).add(c)
        seen[v] = c
    return seen, conflicts


# ── colour awareness ─────────────────────────────────────────────────────────
# Colour must stay a distinguishing dimension: a pair whose alias is one colour but whose
# canonical is a DIFFERENT colour (e.g. v-plo+Pk -> v-plo+Gy) would collapse pink into grey
# and is SKIPPED. Same-colour normalisations and non-colour pairs (6337-P2 -> BD-6337-P2)
# are kept. The colour is the SKU's trailing token, normalised via the synonym map below;
# a trailing pack/size code (P2, x4, ...) is NOT a colour. Edit this map if a code is
# mis-classified — the dry-run prints every skipped pair so you can check.
COLOUR_SYNONYMS = {
    "gy": "grey", "gry": "grey", "grey": "grey", "gray": "grey",
    "cgy": "coolgrey", "cgry": "coolgrey", "cgrey": "coolgrey",
    "ch": "charcoal", "chr": "charcoal", "charcoal": "charcoal",
    "sl": "silver", "slv": "silver", "silver": "silver",
    "wt": "white", "wht": "white", "white": "white",
    "cr": "cream", "crm": "cream", "cream": "cream",
    "bk": "black", "blk": "black", "black": "black",
    "bl": "blue", "blue": "blue",
    "nbl": "navy", "nb": "navy", "nvy": "navy", "navy": "navy",
    "pk": "pink", "pnk": "pink", "pink": "pink",
    "pu": "purple", "ppl": "purple", "purple": "purple",
    "rd": "red", "red": "red", "lc": "lilac", "lilac": "lilac",
    "yl": "yellow", "ylw": "yellow", "yellow": "yellow",
    "br": "brown", "brn": "brown", "brown": "brown",
    "brg": "burgundy", "burg": "burgundy", "burgundy": "burgundy",
    "bg": "beige", "beige": "beige", "sd": "sand", "sand": "sand",
    "tl": "teal", "teal": "teal", "gn": "green", "grn": "green", "green": "green",
    "degg": "duckegg", "duckegg": "duckegg", "mr": "maroon", "maroon": "maroon",
    "gd": "gold", "gold": "gold", "or": "orange", "org": "orange", "orange": "orange",
    "fu": "fuchsia", "fuchsia": "fuchsia",
}


def sku_colour(sku):
    """The SKU's colour family (normalised) from its trailing token, or None if the SKU
    has no recognised colour suffix (e.g. it ends in a pack code like P2)."""
    import re
    toks = [t for t in re.split(r"[-+_ =]+", sku or "") if t]
    if not toks:
        return None
    return COLOUR_SYNONYMS.get(toks[-1].lower())


def is_colour_crossing(variant, canonical):
    """True when the alias carries a colour that the canonical does NOT preserve — i.e.
    applying it would merge one colour into a different colour (or strip the colour)."""
    ca = sku_colour(variant)
    return ca is not None and ca != sku_colour(canonical)


def main():
    apply = "--apply" in sys.argv
    allow_colour = "--allow-colour-merge" in sys.argv   # off by default: colours stay distinct
    xlsx = os.environ.get("SKU_MAP_XLSX", "sku_map_MASTER.xlsx")
    ordered, n_rows = load_pairs(xlsx)
    pairs, conflicts = dedup(ordered)

    # Colour guard: drop pairs that would merge one colour into another (unless overridden).
    keep_pairs, skipped_colour = {}, []
    for v, c in pairs.items():
        if (not allow_colour) and is_colour_crossing(v, c):
            skipped_colour.append((v, c, sku_colour(v), sku_colour(c)))
        else:
            keep_pairs[v] = c

    print(f"\n== sku_map_MASTER sync ==  mode: {'APPLY' if apply else 'DRY-RUN'}"
          f"{'  (colour-merge ALLOWED)' if allow_colour else '  (colours kept distinct)'}")
    print(f"   xlsx : {xlsx}   sheet '{SHEET}'   canonical col {CANON_HDR!r}")
    print(f"   {n_rows} data rows -> {len(ordered)} alias cells -> {len(pairs)} unique variant_sku pairs")
    print(f"   colour-crossing pairs SKIPPED: {len(skipped_colour)}  ->  {len(keep_pairs)} pairs will be applied")
    if skipped_colour:
        print("   skipped (alias colour -> canonical colour would collapse — kept separate):")
        for v, c, ca, cc in sorted(skipped_colour):
            print(f"       {v!r} [{ca}]  ✗->  {c!r} [{cc or 'no-colour'}]")
    if conflicts:
        print(f"   ⚠ {len(conflicts)} variant_sku(s) map to >1 canonical in the sheet (LAST wins — fix in the sheet):")
        for v, cs in list(conflicts.items())[:12]:
            print(f"       {v!r} -> {sorted(cs)}  (using {pairs[v]!r})")

    conn = db.connect()
    print(f"   DB   : {'Postgres (LIVE)' if db.is_postgres() else 'local SQLite'}")

    # 1) alias upsert delta vs current cogs_aliases ---------------------------------
    cur = {r["variant_sku"]: r["canonical_sku"]
           for r in conn.execute("SELECT variant_sku, canonical_sku FROM cogs_aliases").fetchall()}
    new_ct = sum(1 for v in keep_pairs if v not in cur)
    chg_ct = sum(1 for v, c in keep_pairs.items() if v in cur and cur[v] != c)
    same_ct = sum(1 for v, c in keep_pairs.items() if cur.get(v) == c)
    print("\n 1) ALIAS IMPORT into cogs_aliases:")
    print(f"      new variant_skus   : {new_ct}")
    print(f"      changed canonical  : {chg_ct}")
    print(f"      already correct    : {same_ct}")
    print(f"      -> to upsert       : {new_ct + chg_ct} of {len(keep_pairs)}")
    if new_ct + chg_ct == 0 and same_ct == 0:
        print("      ⚠ zero matches — you're likely reading the wrong column/sheet.")
    # existing colour-crossing aliases already in the table (a prior seed may have added some)
    existing_cross = [(v, c) for v, c in cur.items() if is_colour_crossing(v, c)]
    if existing_cross and not allow_colour:
        print(f"      ⚠ {len(existing_cross)} colour-crossing alias(es) ALREADY in cogs_aliases (this import "
              f"leaves them as-is; tell me if you want them removed so those colours split back out):")
        for v, c in sorted(existing_cross)[:12]:
            print(f"          {v!r} [{sku_colour(v)}] -> {c!r} [{sku_colour(c) or 'no-colour'}]")

    # 2) projected re-map of pl_line_items (READ-ONLY simulation) --------------------
    new_map = dict(cur)
    new_map.update(keep_pairs)

    def resolve(sku):
        if not sku:
            return None
        norm = html.unescape(sku).strip()       # mirrors resolve_to_canonical exactly
        return new_map.get(norm, norm)

    li = conn.execute("SELECT sku, canonical_sku FROM pl_line_items").fetchall()
    total_rows = len(li)
    cur_canon, new_canon = set(), set()
    changed_rows = 0
    merges = {}   # (old, new) -> row count
    for r in li:
        old = r["canonical_sku"]
        new = resolve(r["sku"])
        cur_canon.add(old)
        new_canon.add(new)
        if new != old:
            changed_rows += 1
            merges[(old, new)] = merges.get((old, new), 0) + 1
    print(f"\n 2) RE-MAP pl_line_items ({total_rows} rows):")
    print(f"      rows changing canonical    : {changed_rows}")
    print(f"      distinct canonicals BEFORE : {len(cur_canon)}")
    print(f"      distinct canonicals AFTER  : {len(new_canon)}   (−{len(cur_canon) - len(new_canon)})")
    if merges:
        print("      top changes (old -> new : rows):")
        for (old, new), n in sorted(merges.items(), key=lambda x: -x[1])[:30]:
            print(f"        {str(old):<18} -> {str(new):<18} : {n}")

    if not apply:
        print("\n   DRY-RUN only — nothing written. Re-run with --apply AFTER a DB backup.\n")
        conn.close()
        return

    # ---- APPLY -------------------------------------------------------------------
    print("\n APPLYING…")
    ts = _now()
    for v, c in keep_pairs.items():
        conn.execute(
            "INSERT INTO cogs_aliases (variant_sku, canonical_sku, source, updated_at) "
            "VALUES (?,?,?,?) ON CONFLICT(variant_sku) DO UPDATE SET "
            "canonical_sku=excluded.canonical_sku, source=excluded.source, updated_at=excluded.updated_at",
            (v, c, "xlsx_master", ts))
    conn.commit()
    print(f"   upserted {len(keep_pairs)} alias pairs into cogs_aliases ({new_ct} new, {chg_ct} changed; "
          f"{len(skipped_colour)} colour-crossing pairs skipped).")
    before_distinct = conn.execute(
        "SELECT COUNT(DISTINCT canonical_sku) AS c FROM pl_line_items").fetchone()["c"]
    conn.close()

    import pl_tracker
    print("   reprocessing ALL pl_line_items from stored events (no network; may take minutes)…")
    n = pl_tracker.reprocess_from_stored_events()
    print(f"   reprocessed {n} line item(s).")

    conn = db.connect()
    after_distinct = conn.execute(
        "SELECT COUNT(DISTINCT canonical_sku) AS c FROM pl_line_items").fetchone()["c"]
    conn.close()
    print(f"\n   distinct canonicals: {before_distinct} -> {after_distinct}  "
          f"(−{before_distinct - after_distinct})")
    print("   Done. Reload /pl — split rows should now be merged.\n")


if __name__ == "__main__":
    main()
